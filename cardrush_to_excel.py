from pathlib import Path
# -*- coding: utf-8 -*-
r"""
CardRush（デュエル・マスターズ）買取表 → Excelファイル直書き込み（堅牢版・Playwright対応）
- driver選択: requests / playwright（--driver で切替）
- is_hot フィルタはデフォルトOFF（--hot-only でON）
- (name, model_number) で重複排除（(虹)/(虹アイコン) を同一扱いの正規化名で判定）
- レコード末尾に『比較用カード名（正規化名）』を付与
- リトライ & ジッター付きスリープ、UAローテーション、Cookie取得
- .xls は Windows + Excel(pywin32) で .xlsm に自動変換 → openpyxl で書き込み
- 取得0件/しきい値未満は終了コード2で失敗（--min-rows）
- パスはNFC正規化、保存後のサイズ/mtimeをログ出力（--debug）
"""

import argparse
import json
import os
import re
import time
import random
import unicodedata
from typing import Any, Dict, Iterable, List, Tuple, Optional, Set

import requests
from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from urllib.parse import urlencode

from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException

# ---- Playwright (任意) ----
try:
    from playwright.sync_api import sync_playwright
except Exception:
    sync_playwright = None

BASE_URL = "https://cardrush.media/duel_masters/buying_prices"
DEFAULT_LIMIT = 120
DEFAULT_MAX_PAGES = 200
DEFAULT_SLEEP_MS = 400

# ベースUA
UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/118.0.0.0 Safari/537.36"
)

# ローテーション用UA
UAS = [
    UA,
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
]

HEADERS_BASE = {
    "Accept": "text/html,application/xhtml+xml,application/json;q=0.9,*/*;q=0.8",
    "Accept-Language": "ja,en;q=0.8",
    "Referer": BASE_URL,
    "X-Requested-With": "XMLHttpRequest",
}

# ===== 名前正規化（(虹) ≒ (虹アイコン) を同一扱い） =====
_ICON_EQUIV_PATTERNS = [
    re.compile(r'[\(\（]\s*虹\s*アイコン\s*[\)\）]'),
    re.compile(r'[\(\（]\s*虹\s*[\)\）]'),
]
_WS = re.compile(r'\s+')

def _zen2han_keep_kana(s: str) -> str:
    res = []
    for ch in s:
        code = ord(ch)
        if 0xFF01 <= code <= 0xFF5E:
            res.append(chr(code - 0xFEE0))
        else:
            res.append(ch)
    return ''.join(res)

def canon_icon_tags(name: str) -> str:
    if not name:
        return ""
    s = str(name)
    for pat in _ICON_EQUIV_PATTERNS:
        s = pat.sub("", s)
    s = _WS.sub(" ", s).strip()
    return s

def normalize_name_for_match(name: str) -> str:
    if not name:
        return ""
    s = unicodedata.normalize("NFKC", str(name))
    s = _zen2han_keep_kana(s)
    s = canon_icon_tags(s)
    s = s.lower()
    return _WS.sub(" ", s).strip()

def normalize_name_for_dedup(name: str) -> str:
    """
    重複排除用：
    - 大文字小文字、全角半角、余計な空白は正規化
    - (虹) / (虹アイコン) は "消さない"（区別する）
    """
    if not name:
        return ""
    s = unicodedata.normalize("NFKC", str(name))
    s = _zen2han_keep_kana(s)
    s = _WS.sub(" ", s).strip()
    return s.lower()

# ---------- HTTP (requests) ----------
def make_session(ua: Optional[str] = None) -> requests.Session:
    s = requests.Session()
    s.headers.update({**HEADERS_BASE, "User-Agent": ua or random.choice(UAS)})
    # 403はアプリ側で制御するので自動リトライは薄めに
    retry = Retry(
        total=0,
        backoff_factor=0,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset(["GET"]),
        raise_on_status=False,
    )
    s.mount("https://", HTTPAdapter(max_retries=retry))
    return s

def build_params(page: int, limit: int, hot_only: bool) -> List[Tuple[str, str]]:
    params = [
        ("displayMode", "リスト"),
        ("limit", str(limit)),
        ("page", str(page)),
        ("sort[key]", "amount"),
        ("sort[order]", "desc"),
    ]
    if hot_only:
        params.append(("is_hot", "true"))
    return params

# ---------- パース ----------
def normalize_amount(v: Any) -> int:
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = re.sub(r"[^\d]", "", str(v))
    return int(s) if s else 0

def _looks_like_item_dict(d: Dict[str, Any]) -> bool:
    if not isinstance(d, dict):
        return False
    keys = d.keys()
    return any(k in keys for k in ("name", "model_number", "amount"))

def _deep_find_items_array(o: Any) -> List[Dict[str, Any]]:
    best: Optional[List[Dict[str, Any]]] = None
    def dfs(x: Any):
        nonlocal best
        if best is not None:
            return
        if isinstance(x, list) and x:
            dicts = [e for e in x if isinstance(e, dict)]
            if dicts:
                if any(_looks_like_item_dict(dd) for dd in dicts):
                    best = dicts
                    return
                for dd in dicts:
                    dfs(dd)
            else:
                for e in x:
                    dfs(e)
        elif isinstance(x, dict):
            for k in ("buying_prices", "items", "list", "data", "records", "products", "result"):
                if k in x:
                    dfs(x[k])
                    if best is not None:
                        return
            for v in x.values():
                dfs(v)
    dfs(o)
    return best or []

def parse_items_from_json(obj: Any) -> List[Dict[str, Any]]:
    if obj is None:
        return []
    if isinstance(obj, list):
        dicts = [e for e in obj if isinstance(e, dict)]
        if dicts:
            if any(_looks_like_item_dict(dd) for dd in dicts):
                return dicts
            return _deep_find_items_array(obj)
        return _deep_find_items_array(obj)
    if isinstance(obj, dict):
        for k in ("buying_prices", "items", "list", "data", "records", "products", "result"):
            if isinstance(obj.get(k), list):
                return parse_items_from_json(obj[k])
        return _deep_find_items_array(obj)
    return []

def parse_items_from_next_data(html_text: str) -> List[Dict[str, Any]]:
    soup = BeautifulSoup(html_text, "html.parser")
    tag = soup.find("script", id="__NEXT_DATA__")
    if not tag or not tag.text:
        return []
    try:
        data = json.loads(tag.text)
    except Exception:
        return []
    return parse_items_from_json(data)

# ---------- 1ページ取得（requests + 再試行） ----------
def fetch_page_requests(page: int, limit: int, hot_only: bool, timeout: float = 20.0, debug: bool=False):
    for attempt in range(1, 6):  # 最大5回、指数バックオフ
        ua = random.choice(UAS)
        sess = make_session(ua)
        # 事前にトップを踏んでCookie獲得（失敗しても続行）
        try:
            sess.get(BASE_URL, timeout=timeout)
        except Exception:
            pass

        params = build_params(page, limit, hot_only)
        try:
            r = sess.get(BASE_URL, params=params, timeout=timeout)
        except requests.RequestException as e:
            if debug:
                print(f"[debug] request error attempt={attempt}: {e}")
            time.sleep(1 + attempt * 0.5)
            continue

        used_url = r.url
        txt = (r.text or "").strip()
        status = r.status_code
        if debug:
            print(f"[debug] fetch_page(req) try={attempt} ua={ua[:40]}... page={page} status={status} len={len(txt)} url={used_url}")

        if status == 403:
            backoff = (2 ** attempt) + random.uniform(0.0, 1.0)
            time.sleep(backoff)
            continue

        if txt.startswith("{") or txt.startswith("[" ):
            try:
                obj = r.json()
                items = parse_items_from_json(obj)
                if items:
                    return items, used_url, status
            except Exception as e:
                if debug:
                    print(f"[debug] json parse error: {e}")

        items = parse_items_from_next_data(txt)
        return items, used_url, status

    if debug:
        print("[debug] fetch_page(req): all retries exhausted (treat as 403)")
    return [], BASE_URL, 403

# ---------- 1ページ取得（Playwright） ----------
def fetch_page_playwright(page: int, limit: int, hot_only: bool, timeout: float = 25.0, debug: bool=False):
    if sync_playwright is None:
        raise RuntimeError("Playwright is not installed. Run: pip install playwright && python -m playwright install chromium")
    params = build_params(page, limit, hot_only)
    url = f"{BASE_URL}?{urlencode(params, doseq=True)}"
    ua = random.choice(UAS)
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(user_agent=ua, locale="ja-JP")
        pageobj = context.new_page()
        pageobj.set_default_timeout(int(timeout * 1000))

        # まずトップでCookie
        try:
            pageobj.goto(BASE_URL, wait_until="domcontentloaded")
        except Exception:
            pass

        resp = pageobj.goto(url, wait_until="networkidle")
        status = resp.status if resp else 200
        if debug:
            print(f"[debug] fetch_page(pw) status={status} url={url}")

        items: List[Dict[str, Any]] = []
        try:
            pageobj.wait_for_selector("script#__NEXT_DATA__", timeout=int(timeout * 1000))
            txt = pageobj.locator("script#__NEXT_DATA__").inner_text()
            data = json.loads(txt)
            items = parse_items_from_json(data)
        except Exception:
            html = pageobj.content()
            items = parse_items_from_next_data(html)
        finally:
            context.close()
            browser.close()
        return items, url, status

# ---------- レコード整形 ----------
def items_to_rows(items: Iterable[Dict[str, Any]], used_url: str) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for it in items:
        if not isinstance(it, dict):
            continue
        name = it.get("name", "")
        model = it.get("model_number", "")
        amount = normalize_amount(it.get("amount"))
        rarity = it.get("rarity", "")
        category = it.get("display_category", "")

        name_norm = normalize_name_for_match(name)

        if name or model or amount:
            rows.append([name, model, amount, category, rarity, used_url, name_norm])
    return rows

# ---------- 全ページ取得 ----------
def scrape_all(limit=DEFAULT_LIMIT, max_pages=DEFAULT_MAX_PAGES, sleep_ms=DEFAULT_SLEEP_MS,
               hot_only=False, debug: bool=False, driver: str="requests"):
    all_rows: List[List[Any]] = []
    seen: Set[Tuple[str, str]] = set()
    fetcher = fetch_page_playwright if driver == "playwright" else fetch_page_requests

    for page in range(1, max_pages + 1):
        items, used_url, status = fetcher(page, limit, hot_only, debug=debug)
        if status >= 400 and not items:
            if debug:
                print(f"[debug] stop: HTTP {status} with no items on page {page}")
            break
        if not items:
            if debug:
                print(f"[debug] stop: no items on page {page}")
            break

        page_rows = items_to_rows(items, used_url)

        # 重複排除（正規化名 + 型番）
        deduped = []
        for r in page_rows:
            orig_name, model = str(r[0]).strip(), str(r[1]).strip()
            key = (normalize_name_for_dedup(orig_name), model)
            if key in seen:
                continue
            seen.add(key)
            deduped.append(r)

        all_rows.extend(deduped)

        if debug:
            print(f"[debug] page {page}: items={len(items)} rows_kept={len(deduped)} cumulative={len(all_rows)}")

        if len(items) < limit:
            if debug:
                print(f"[debug] stop: items<{limit} (last page={page})")
            break

        jitter = random.uniform(0.8, 1.4)
        time.sleep(max(0, sleep_ms) * jitter / 1000.0)
    return all_rows

# ---------- Excelユーティリティ ----------
def _ensure_parent_dir(path: str):
    parent = os.path.dirname(path)
    if parent and not os.path.exists(parent):
        os.makedirs(parent, exist_ok=True)

def _convert_to_xlsm_if_needed(file_path: str) -> str:
    base, _ = os.path.splitext(file_path)
    target_path = base + ".xlsm"
    try:
        import win32com.client as win32
    except Exception:
        raise RuntimeError(
            "旧形式(.xls)を検出しましたが pywin32 が未インストールのため変換できません。\n"
            "Windows + Excel 環境で `pip install pywin32` を実行するか、手動で .xlsm へ保存してください。"
        )
    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False
    try:
        wb = excel.Workbooks.Open(os.path.abspath(file_path))
        wb.SaveAs(os.path.abspath(target_path), 52)  # 52 = xlOpenXMLWorkbookMacroEnabled (.xlsm)
        wb.Close(SaveChanges=False)
    finally:
        excel.Quit()
    return target_path

def _open_or_create_workbook(file_path: str, debug: bool=False):
    file_path = unicodedata.normalize("NFC", file_path)
    _ensure_parent_dir(file_path)
    if not os.path.exists(file_path):
        if debug:
            print(f"[debug] workbook not found, creating new: {file_path}")
        wb = Workbook()
        # 既定の "Sheet" を消して空に
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            wb.remove(wb.active)
        wb.save(file_path)
        return wb, file_path
    try:
        wb = load_workbook(file_path, keep_vba=True)
        return wb, file_path
    except InvalidFileException:
        print("[info] 旧形式/不正形式の可能性あり。Excelで .xlsm に自動変換します...")
        new_path = _convert_to_xlsm_if_needed(file_path)
        print(f"[info] 変換完了: {new_path}（再オープン）")
        wb = load_workbook(new_path, keep_vba=True)
        return wb, new_path

# ---------- Excel書き込み ----------
def write_to_excel(rows: List[List[Any]], file_path: str, sheet_name: str, debug: bool=False) -> Tuple[str, int]:
    if not rows:
        print("[warn] fetched 0 rows. keep previous data (no overwrite).")
        return unicodedata.normalize("NFC", file_path), 0

    wb, actual_path = _open_or_create_workbook(file_path, debug=debug)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if debug:
            print(f"[debug] clear existing sheet: {sheet_name} (rows={ws.max_row})")
        ws.delete_rows(1, ws.max_row)
    else:
        if debug:
            print(f"[debug] create new sheet: {sheet_name} / existing={wb.sheetnames}")
        ws = wb.create_sheet(sheet_name)

    header = ["カード名", "型番", "買取金額(円)", "カテゴリ", "レア", "取得元URL", "比較用カード名"]
    ws.append(header)
    for r in rows:
        ws.append(r)

    wb.save(actual_path)

    try:
        st = os.stat(actual_path)
        print(f"[info] saved: {os.path.abspath(actual_path)} size={st.st_size} mtime={time.strftime('%Y-%m-%d %H:%M:%S', time.gmtime(st.st_mtime))}")
    except Exception:
        pass

    updated_cells = len(rows) * len(header)
    print(f"[info] rows_written={len(rows)} updated_cells~={updated_cells}")
    return actual_path, len(rows)

# ---------- CLI ----------
def main():
    ap = argparse.ArgumentParser(description="CardRush（DM）→ Excelファイル直書き（全件取得・Playwright対応）")
    ap.add_argument("--file-path",
        default=str(Path(__file__).resolve().parent / "buylist.xlsm"),
        help="出力先Excelファイルのパス（.xlsm 推奨。旧 .xls は自動変換）")
    ap.add_argument("--sheet-name", default="シート2", help="出力先シート名")
    ap.add_argument("--limit", type=int, default=DEFAULT_LIMIT)
    ap.add_argument("--max-pages", type=int, default=DEFAULT_MAX_PAGES)
    ap.add_argument("--sleep-ms", type=int, default=DEFAULT_SLEEP_MS)
    ap.add_argument("--hot-only", action="store_true")
    ap.add_argument("--min-rows", type=int, default=0, help="最小許容件数（未満なら終了コード2で落とす）")
    ap.add_argument("--driver", choices=["requests", "playwright"], default="requests",
                    help="取得方式（requests / playwright）")
    ap.add_argument("--debug", action="store_true")
    args = ap.parse_args()

    if args.debug:
        print(f"[debug] start with args: {vars(args)}")

    rows = scrape_all(
        limit=args.limit,
        max_pages=args.max_pages,
        sleep_ms=args.sleep_ms,
        hot_only=args.hot_only,
        debug=args.debug,
        driver=args.driver
    )

    if len(rows) == 0:
        print("[err] scraped 0 rows (blocked or parse failed).")
        raise SystemExit(2)

    if args.min_rows and len(rows) < args.min_rows:
        print(f"[err] too few rows: {len(rows)} (<{args.min_rows}).")
        raise SystemExit(2)

    path, n = write_to_excel(rows, args.file_path, args.sheet_name, debug=args.debug)
    print(f"[OK] {n} 件を書き込み完了: {args.sheet_name} → {path}")

if __name__ == "__main__":
    main()


