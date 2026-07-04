from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime, date
import pandas as pd

# ========= 設定 =========
excel_path = Path(__file__).resolve().parent / "buylist.xlsm"
sheet_name = "シート1"
csv_path = excel_path.parent / "Mycaアップロード用CSV.csv"

# Excelで日付化されたくない列名を入れる
# 例: ["カード番号", "型番", "品番"]
force_text_columns = []
# 全列で日付っぽい値を保護したいなら True
protect_all_date_like_values = True
# ========================

print("Excel:", excel_path)
print("Sheet:", sheet_name)
print("CSV :", csv_path)

wb = load_workbook(excel_path, data_only=True)
ws = wb[sheet_name]

rows = list(ws.iter_rows(values_only=False))
if not rows:
    raise ValueError("シートが空です")

def is_date_like_string(s: str) -> bool:
    s = str(s).strip()
    parts = s.split("/")
    if len(parts) != 2:
        return False
    a, b = parts
    return a.isdigit() and b.isdigit()

def cell_raw_text(cell):
    v = cell.value
    if v is None:
        return ""

    # Excel内部で日付として持っている場合は m/d に戻す
    if isinstance(v, (datetime, date)):
        return f"{v.month}/{v.day}"

    return str(v).strip()

# ヘッダー
header = [cell_raw_text(c) for c in rows[0]]

body = []
for row in rows[1:]:
    out = []
    for idx, cell in enumerate(row):
        value = cell_raw_text(cell)
        col_name = header[idx] if idx < len(header) else f"col_{idx}"

        should_protect = False
        if value and is_date_like_string(value):
            if protect_all_date_like_values or col_name in force_text_columns:
                should_protect = True

        # Excelで開いた時に 1/1 → 日付化されるのを防ぐ
        if should_protect:
            value = f'="{value}"'

        out.append(value)
    body.append(out)

df = pd.DataFrame(body, columns=header)

# 空行削除
df = df.replace("", pd.NA).dropna(how="all").fillna("")

df.to_csv(
    csv_path,
    index=False,
    encoding="utf-8-sig"
)

print("CSV作成完了:", csv_path)
print("行数:", len(df))
