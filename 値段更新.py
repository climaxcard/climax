# -*- coding: utf-8 -*-
r"""
buylist.xlsm：シート1とシート2をマッチングし、
マッチしたカードのみ「シート1のO列（買取価格）」を上書き更新する。

- マッチングシステムは現状コードを維持（STRICT→LOOSE + 名前ガード + 秘一致 + 虹優先 等）
- 差分比較シート出力は一切しない

価格更新ルール:
- new = floor(S2_price * 0.95)
- 丸め:
    * new < 10000  : 100円単位で切り捨て（285 -> 200）
    * new >= 10000 : 1000円単位で切り捨て（10950 -> 10000）
- マッチしない行は O列を変更しない
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import unicodedata
import re
from difflib import SequenceMatcher
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


# ===== 設定 =====
XLSM_PATH = str(Path(__file__).resolve().parent / "buylist.xlsm")

SHEET1 = "シート1"
SHEET2 = "シート2"

# 列（1始まり）
# シート1
S1_COL_NAME  = 3   # C
S1_COL_E     = 5   # E
S1_COL_F     = 6   # F
S1_COL_PRICE = 15  # O
S1_COL_LOCK  = 16
S1_HEADER_ROWS = 1

# シート2
S2_COL_NAME  = 1   # A
S2_COL_MODEL = 2   # B
S2_COL_PRICE = 3   # C
S2_HEADER_ROWS = 1

# ===== 名前ガード =====
NAME_HARD_REJECT = 0.75
NAME_REQUIRED_FOR_STRICT = 0.80
NAME_REQUIRED_FOR_LOOSE = 0.84

# ===== 正規化 =====
_re_spaces = re.compile(r"\s+")
_re_keep_alnum = re.compile(r"[^0-9A-Z]+")   # 英数字以外を除去（LOOSE用）
_re_drop_ab_lower = re.compile(r"[ab]")      # 小文字 a/b のみ削除（A/Bは消さない）

# ===== 名前比較用：スコアを落としやすい記号を除去/統一（比較だけ）=====
_NAME_PUNCT_TO_DROP = str.maketrans({
    "!": "", "！": "",
    "?": "", "？": "",
    "·": "", "・": "",
    "’": "", "'": "", "“": "", "”": "", '"': "",
    ",": "", "，": "", ".": "", "．": "",
    "‐": "", "-": "", "‒": "", "–": "", "—": "", "―": "", "−": "",
    "／": "/",  # 統一
    "：": ":",  # 統一
})


# ===== セル値 → 安定文字列（★日付型＋日付シリアル対策）=====
def cell_to_text(v: Any) -> str:
    """
    型番セル値を安定して文字列化する。

    - datetime/date は必ず M/D/YYYY
    - int/float も「日付っぽい範囲」の場合は Excel日付シリアルとして解釈し、M/D/YYYY
    """
    if v is None:
        return ""
    if isinstance(v, datetime):
        d = v.date()
        return f"{d.month}/{d.day}/{d.year}"
    if isinstance(v, date):
        return f"{v.month}/{v.day}/{v.year}"

    # 日付シリアルっぽい数値を救済（だいたい 2000-01-01 ~ 2100-12-31）
    if isinstance(v, (int, float)):
        if 30000 <= float(v) <= 80000:
            try:
                d = from_excel(v)  # datetime
                dd = d.date() if isinstance(d, datetime) else d
                return f"{dd.month}/{dd.day}/{dd.year}"
            except Exception:
                pass
        if float(v).is_integer():
            return str(int(v))
        return str(v)

    return str(v).strip()


# ===== ローマ数字（Unicode→ASCII）=====
_ROMAN_MAP = {
    "Ⅰ": "I", "Ⅱ": "II", "Ⅲ": "III", "Ⅳ": "IV", "Ⅴ": "V",
    "Ⅵ": "VI", "Ⅶ": "VII", "Ⅷ": "VIII", "Ⅸ": "IX", "Ⅹ": "X",
    "Ⅺ": "XI", "Ⅻ": "XII", "Ⅼ": "L", "Ⅽ": "C", "Ⅾ": "D", "Ⅿ": "M",
    "ⅰ": "I", "ⅱ": "II", "ⅲ": "III", "ⅳ": "IV", "ⅴ": "V",
    "ⅵ": "VI", "ⅶ": "VII", "ⅷ": "VIII", "ⅸ": "IX", "ⅹ": "X",
    "ⅺ": "XI", "ⅻ": "XII", "ⅼ": "L", "ⅽ": "C", "ⅾ": "D", "ⅿ": "M",
}

def roman_unicode_to_ascii(s: str) -> str:
    if not s:
        return s
    for k in sorted(_ROMAN_MAP.keys(), key=len, reverse=True):
        s = s.replace(k, _ROMAN_MAP[k])
    return s

def nfkc(s: str) -> str:
    s = unicodedata.normalize("NFKC", s)
    s = roman_unicode_to_ascii(s)
    return s

def nfkc_upper(s: str) -> str:
    return nfkc(s).upper()

def secret_rank(s: str) -> str:
    """
    型番の秘種別を判定する。
    STD超秘1 と STD秘1 を別物として扱うため、boolではなく種類で返す。

    return:
      "CHOHI"  = 超秘
      "HI"     = 秘
      "NONE"   = 秘なし
    """
    if not s:
        return "NONE"

    s = nfkc(str(s))

    if "超秘" in s:
        return "CHOHI"

    if "秘" in s:
        return "HI"

    return "NONE"


# ===== 虹フラグ =====
def has_rainbow_mark_raw(name_raw: str) -> bool:
    if not name_raw:
        return False
    s = nfkc(name_raw)
    return ("(虹)" in s) or ("（虹）" in s) or ("(虹アイコン)" in s) or ("（虹アイコン）" in s)


# ===== S2型番トリム（2つ目の/以降を無視、ただし /Y は例外、日付は例外）=====
_re_date_mdy = re.compile(r"^\s*\d{1,2}/\d{1,2}/\d{2,4}\s*$")
_re_contains_slash_y = re.compile(r"/\s*Y", flags=re.IGNORECASE)

def trim_after_second_slash_with_y_exception(s: str) -> str:
    if not s:
        return s
    s0 = s
    s_chk = nfkc(s0)

    if _re_date_mdy.match(s_chk):
        return s0
    if _re_contains_slash_y.search(s_chk):
        return s0

    if s_chk.count("/") >= 2:
        parts = s0.split("/")
        if len(parts) >= 2:
            return parts[0] + "/" + parts[1]
    return s0


# ===== DM処理（E列キー作成にのみ適用）=====
_re_dmrp = re.compile(r"DMRP", flags=re.IGNORECASE)
_re_dmex = re.compile(r"DMEX", flags=re.IGNORECASE)
_re_dm_2223242526 = re.compile(r"DM(?=(22|23|24|25|26))", flags=re.IGNORECASE)
_re_dm_before_letter = re.compile(r"DM(?=[A-Z])", flags=re.IGNORECASE)
_re_has_slash_y = re.compile(r"/\s*Y", flags=re.IGNORECASE)

def apply_dm_rule_on_e_for_key(e_raw: str, *, keep_dm_2223242526: bool = False) -> str:
    if not e_raw:
        return e_raw

    s = e_raw
    s = _re_dmrp.sub("RP", s)    # DMRP -> RP
    s = _re_dmex.sub("EX", s)    # DMEX -> EX

    if not keep_dm_2223242526:
        s = _re_dm_2223242526.sub("", s)  # DM22/23/24/25/26 -> 22/23/24/25/26

    def _drop_dm_if_letter_not_rcxd(m: re.Match) -> str:
        idx = m.end()
        nxt = s[idx:idx+1].upper()
        if nxt in ("R", "C", "X", "D"):
            return "DM"
        return ""

    s = _re_dm_before_letter.sub(_drop_dm_if_letter_not_rcxd, s)
    return s


# ===== raw完全一致用（記号・カッコ含めて一致させたい用途）=====
def norm_name_raw_exact(s: str) -> str:
    if not s:
        return ""
    return nfkc_upper(str(s).strip())


# ===== 名前正規化（比較用）=====
def norm_name(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""

    s = nfkc_upper(s)

    # 比較用は虹表記を除去（候補選択で虹フラグ制御）
    s = s.replace("(虹アイコン)", "").replace("（虹アイコン）", "")
    s = s.replace("(虹)", "").replace("（虹）", "")

    # 比較だけ：記号除去/統一
    s = s.translate(_NAME_PUNCT_TO_DROP)
    s = _re_spaces.sub("", s)

    # カッコ文字だけ除去（中身は残る）
    for ch in ["【", "】", "[", "]", "(", ")", "（", "）"]:
        s = s.replace(ch, "")
    return s


# ===== 型番キー正規化（共通）=====
def _apply_model_common_cleanup_for_key(s: str) -> str:
    if not s:
        return s
    s = nfkc(s)
    s = _re_drop_ab_lower.sub("", s)  # 小文字a/bだけ削除
    s = s.upper()
    return s

def model_strict(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    s = _apply_model_common_cleanup_for_key(s)
    s = _re_spaces.sub("", s)
    return s

def model_loose(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    s = _apply_model_common_cleanup_for_key(s)
    s = _re_keep_alnum.sub("", s)
    return s


# ===== S1 表示/キー =====
def s1_model_raw_f_display(f: Any) -> str:
    raw = cell_to_text(f)
    return nfkc(raw).strip()

def s1_model_ef_for_key(e: Any, f: Any) -> str:
    ee = nfkc(cell_to_text(e))
    ff = nfkc(cell_to_text(f))

    # F側が /Y を含むなら DM22/23/24/25/26 の “DM削除” をしない
    keep_dm = bool(_re_has_slash_y.search(nfkc_upper(ff)))
    ee = apply_dm_rule_on_e_for_key(ee, keep_dm_2223242526=keep_dm)  # E列のみ
    return f"{ee}{ff}".strip()


# ===== S2 表示用 raw =====
def s2_model_raw_display(model_raw: Any) -> str:
    raw = cell_to_text(model_raw)
    raw = trim_after_second_slash_with_y_exception(raw)
    raw = nfkc(raw)
    raw = _re_drop_ab_lower.sub("", raw)  # 表示でも小文字a/bだけ消す
    return raw.strip()


# ===== 数値変換 =====
def to_number(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None

def ratio(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()


# ===== S2 index =====
# rows[r] = (model_disp, name_raw, price, name_norm, rainbow_flag, name_raw_exact)
def build_s2_indexes(ws2):
    strict_map: Dict[str, List[int]] = {}
    loose_map: Dict[str, List[int]] = {}
    rows: Dict[int, Tuple[str, str, Optional[float], str, bool, str]] = {}

    for r in range(S2_HEADER_ROWS + 1, ws2.max_row + 1):
        name_raw = ws2.cell(r, S2_COL_NAME).value
        model_raw = ws2.cell(r, S2_COL_MODEL).value
        price = to_number(ws2.cell(r, S2_COL_PRICE).value)

        name_raw_s = "" if name_raw is None else str(name_raw).strip()
        model_disp = s2_model_raw_display(model_raw)

        k_strict = model_strict(model_disp)
        k_loose = model_loose(model_disp)

        n_name = norm_name(name_raw_s)
        rb = has_rainbow_mark_raw(name_raw_s)
        raw_exact = norm_name_raw_exact(name_raw_s)

        rows[r] = (model_disp, name_raw_s, price, n_name, rb, raw_exact)

        if k_strict:
            strict_map.setdefault(k_strict, []).append(r)
        if k_loose:
            loose_map.setdefault(k_loose, []).append(r)

    return strict_map, loose_map, rows


def pick_best_by_name_with_rainbow_preference(
    candidates: List[int],
    s1_name_n: str,
    s1_name_exact: str,
    s1_rainbow: bool,
    s2_rows: Dict[int, Tuple[str, str, Optional[float], str, bool, str]],
) -> Tuple[Optional[int], float]:
    """
    虹優先ロジック + raw完全一致優先 + 正規化名完全一致優先 + 類似度ベスト
    """
    def filter_by_rainbow_rule(cands: List[int]) -> List[int]:
        if not cands:
            return []
        if s1_rainbow:
            return [r2 for r2 in cands if s2_rows[r2][4] is True]
        non_rb = [r2 for r2 in cands if s2_rows[r2][4] is False]
        return non_rb if non_rb else cands

    def best_of(cands: List[int]) -> Tuple[Optional[int], float]:
        if not cands:
            return None, 0.0

        exact_raw = [r2 for r2 in cands if s2_rows[r2][5] == s1_name_exact]
        if exact_raw:
            return exact_raw[0], 1.0

        exact_norm = [r2 for r2 in cands if s2_rows[r2][3] == s1_name_n]
        if exact_norm:
            return exact_norm[0], 1.0

        best_r = None
        best_sc = 0.0
        for r2 in cands:
            sc = ratio(s1_name_n, s2_rows[r2][3])
            if sc > best_sc:
                best_sc = sc
                best_r = r2
        return best_r, best_sc

    c2 = filter_by_rainbow_rule(candidates)
    return best_of(c2)


# ===== 価格計算（要望の丸め）=====
def calc_new_price_from_s2(s2_price: float) -> int:
    """
    new = floor(s2_price * 0.95)

    例外:
      - S2=10  → 10
      - S2=30  → 10
      - S2=50  → 10
      - S2=100 → 50

    丸め:
      - base >= 100000 : 10000円単位で切り上げ（127000 -> 130000）
      - base >= 10000  : 1000円単位で切り捨て（10950 -> 10000）
      - base < 10000   : 100円単位で切り捨て（285 -> 200）
    """
    s2_int = int(round(float(s2_price)))

    # 低価格例外
    if s2_int in (10, 30, 50):
        return 10
    if s2_int == 100:
        return 50

    base = int((s2_int * 95) // 100)

    # ★追加：10万以上は1万円単位で切り上げ
    if base >= 100000:
        return ((base + 9999) // 10000) * 10000

    # 既存：1万以上は1000円単位で切り捨て
    if base >= 10000:
        return (base // 1000) * 1000

    # 既存：1万未満は100円単位で切り捨て
    return (base // 100) * 100

def is_checked_cell(v) -> bool:
    """
    Excelのチェック列(P列)が「チェックON」かどうかを判定する。
    TRUE/True/1/✓/✔/☑ などをON扱いにする。
    """
    if v is None:
        return False

    if isinstance(v, bool):
        return v

    if isinstance(v, (int, float)):
        return v != 0

    s = str(v).strip()
    if s == "":
        return False

    s_upper = s.upper()

    if s_upper in {"FALSE", "0", "OFF", "NO"}:
        return False
    if s in {"☐"}:
        return False

    if s_upper in {"TRUE", "1", "ON", "YES"}:
        return True
    if any(mark in s for mark in ["☑", "✔", "✓", "■", "●", "レ"]):
        return True

    return False


def main() -> int:
    xlsm_path = Path(XLSM_PATH)
    if not xlsm_path.exists():
        raise FileNotFoundError(f"見つかりません: {xlsm_path}")

    wb = load_workbook(xlsm_path, keep_vba=True, data_only=False)

    if SHEET1 not in wb.sheetnames:
        raise RuntimeError(f"シートがありません: {SHEET1}")
    if SHEET2 not in wb.sheetnames:
        raise RuntimeError(f"シートがありません: {SHEET2}")

    ws1 = wb[SHEET1]
    ws2 = wb[SHEET2]

    s2_strict, s2_loose, s2_rows = build_s2_indexes(ws2)

    matched = 0
    updated = 0
    skipped_no_s2_price = 0

    for r1 in range(S1_HEADER_ROWS + 1, ws1.max_row + 1):

        # ===== P列ロック（チェックボックス）=====
        lock_flag = ws1.cell(r1, S1_COL_LOCK).value
        if is_checked_cell(lock_flag):
            continue

        s1_name_raw = ws1.cell(r1, S1_COL_NAME).value
        s1_e_raw = ws1.cell(r1, S1_COL_E).value
        s1_f_raw = ws1.cell(r1, S1_COL_F).value

        s1_name = "" if s1_name_raw is None else str(s1_name_raw).strip()
        s1_name_exact = norm_name_raw_exact(s1_name)
        s1_name_n = norm_name(s1_name)
        s1_rainbow = has_rainbow_mark_raw(s1_name)

        s1_model_f_disp = s1_model_raw_f_display(s1_f_raw)
        ef_key_src = s1_model_ef_for_key(s1_e_raw, s1_f_raw)

        ef_strict = model_strict(ef_key_src)
        f_strict  = model_strict(s1_model_f_disp)
        ef_loose  = model_loose(ef_key_src)
        f_loose   = model_loose(s1_model_f_disp)

        # 「秘」種別（STD超秘1 と STD秘1 を別物にする）
        s1_secret_text = nfkc(cell_to_text(s1_e_raw)) + nfkc(cell_to_text(s1_f_raw)) + nfkc(s1_model_f_disp)
        s1_secret_rank = secret_rank(s1_secret_text)

        if not s1_name_n and not ef_strict and not f_strict:
            continue

        r2_final: Optional[int] = None
        name_score = 0.0

        def secret_ok_for_row(r2: int) -> bool:
            m2_disp, _n2_raw, _p2, _n2, _rb2, _exact = s2_rows[r2]
            return s1_secret_rank == secret_rank(m2_disp)

        def pick(cand: List[int]) -> Tuple[Optional[int], float]:
            return pick_best_by_name_with_rainbow_preference(
                cand, s1_name_n, s1_name_exact, s1_rainbow, s2_rows
            )

        # STRICT EF
        if ef_strict and ef_strict in s2_strict:
            cand = [r2 for r2 in s2_strict[ef_strict] if secret_ok_for_row(r2)]
            if cand:
                rr, sc = pick(cand)
                if rr is not None and sc >= NAME_REQUIRED_FOR_STRICT and sc >= NAME_HARD_REJECT:
                    r2_final = rr
                    name_score = sc

        # STRICT F
        if r2_final is None and f_strict and f_strict in s2_strict:
            cand = [r2 for r2 in s2_strict[f_strict] if secret_ok_for_row(r2)]
            if cand:
                rr, sc = pick(cand)
                if rr is not None and sc >= NAME_REQUIRED_FOR_STRICT and sc >= NAME_HARD_REJECT:
                    r2_final = rr
                    name_score = sc

        # LOOSE EF
        if r2_final is None and ef_loose and ef_loose in s2_loose:
            cand = [r2 for r2 in s2_loose[ef_loose] if secret_ok_for_row(r2)]
            if cand:
                rr, sc = pick(cand)
                if rr is not None and sc >= NAME_REQUIRED_FOR_LOOSE and sc >= NAME_HARD_REJECT:
                    r2_final = rr
                    name_score = sc

        # LOOSE F
        if r2_final is None and f_loose and f_loose in s2_loose:
            cand = [r2 for r2 in s2_loose[f_loose] if secret_ok_for_row(r2)]
            if cand:
                rr, sc = pick(cand)
                if rr is not None and sc >= NAME_REQUIRED_FOR_LOOSE and sc >= NAME_HARD_REJECT:
                    r2_final = rr
                    name_score = sc

        # 通常版が虹に吸われるのを止める（既存方針）
        if (r2_final is not None) and (not s1_rainbow):
            _m2_disp, _n2_raw, _p2, _n2_norm, r2_is_rb, _exact = s2_rows[r2_final]
            if r2_is_rb is True:
                exact_name_non_rb = [
                    r2 for r2, (_md, _nr, _pr, n_norm, rb, _ex) in s2_rows.items()
                    if (rb is False) and (n_norm == s1_name_n) and secret_ok_for_row(r2)
                ]
                if exact_name_non_rb:
                    def model_rank(r2: int) -> int:
                        md = s2_rows[r2][0]
                        ks = model_strict(md)
                        kl = model_loose(md)
                        if ks and (ks == ef_strict or ks == f_strict):
                            return 3
                        if kl and (kl == ef_loose or kl == f_loose):
                            return 2
                        return 1

                    r2_final = max(exact_name_non_rb, key=model_rank)
                    name_score = 1.0

        # ===== マッチしたら O列更新 =====
        if r2_final is not None:
            matched += 1
            _m2_disp, _n2_raw, s2_price, _n2, _rb2, _exact = s2_rows[r2_final]

            if s2_price is None:
                skipped_no_s2_price += 1
                continue

            new_price = calc_new_price_from_s2(float(s2_price))
            ws1.cell(r1, S1_COL_PRICE).value = new_price
            updated += 1

    wb.save(xlsm_path)
    wb.close()

    print("✅ 完了: シート2の金額(95%)を丸めて、マッチした行のみシート1(O列)へ上書きしました。")
    print(f"   matched: {matched}")
    print(f"   updated: {updated}")
    print(f"   skipped (S2 price empty): {skipped_no_s2_price}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
